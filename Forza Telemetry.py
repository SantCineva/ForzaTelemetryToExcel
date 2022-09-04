# ####################################################################################
#   ____              _    ____ _                       
#  / ___|  __ _ _ __ | |_ / ___(_)_ __   _____   ____ _ 
#  \___ \ / _` | '_ \| __| |   | | '_ \ / _ \ \ / / _` |
#   ___) | (_| | | | | |_| |___| | | | |  __/\ V / (_| |
#  |____/ \__,_|_| |_|\__|\____|_|_| |_|\___| \_/ \__,_|
#
######################################################################################                                                      

import socket
import struct
import time
import openpyxl
import ForzaHorizonTelemetry as FHT
from openpyxl.utils import get_column_letter


if __name__ == '__main__':

    LogLength = 10                                                                  # Length of time to record
    LogUnit = '%S'                                                                  # Change "%S" to "%M" if you want to set the LogLength in minutes.

    StartTime = int(time.strftime(LogUnit))                                            
    print(f'StartTime is {time.strftime("%H:%M:%S")}')


    OutputFile = 'E:\OutputFile.xlsx'                                               # Output file name
    WorkBook = openpyxl.Workbook()                                                  # Create a new Excel .xlsx document
    WorkSheet = WorkBook.active                                                     # Select the current WorkSheet
    LetterNumber = 1                                                                # Each collumn has a letter assigned, but you can map the letters to numbers
    DataRow = 2                                                                     # The first row of the sheet is occupied by the Header, so the data starts from the second row

    # Titles = ['RaceOn' , 'TimeStamp' , 'MaxRPM' , 'IdleRPM' , 'CurrentRPM' , 'AccelerationX' , 'AccelerationY' , 'AccelerationZ' , 'VelocityX' , 'VelocityY' , 'VelocityZ' , 'AngularVelocityX' , 'AngularVelocityY' , 'AngularVelocityZ' , 'Yaw' , 'Pitch' , 'Roll' , 'NormalizedSuspensionTravelFL' , 'NormalizedSuspensionTravelFR' , 'NormalizedSuspensionTravelRL' , 'NormalizedSuspensionTravelRR' , 'TireSlipRatioFL' , 'TireSlipRatioFR' , 'TireSlipRatioRL' , 'TireSlipRatioRR' , 'WheelRotationSpeedFL' , 'WheelRotationSpeedFR' , 'WheelRotationSpeedRL' , 'WheelRotationSpeedRR' , 'WheelOnRumbleStripFL' , 'WheelOnRumbleStripFR' , 'WheelOnRumbleStripRL' , 'WheelOnRumbleStripRR' , 'WheelInPuddleDepthFL' , 'WheelInPuddleDepthFR' , 'WheelInPuddleDepthRL' , 'WheelInPuddleDepthRR' , 'SurfaceRumbleFrontLeft' , 'SurfaceRumbleFrontRight' , 'SurfaceRumbleRearLeft' , 'SurfaceRumbleRearRight' , 'TireSlipAngleFL' , 'TireSlipAngleFR' , 'TireSlipAngleRL' , 'TireSlipAngleRR' , 'TireCombinedSlipFL' , 'TireCombinedSlipFR' , 'TireCombinedSlipRL' , 'TireCombinedSlipRR' , 'SuspensionTravelMetersFL' , 'SuspensionTravelMetersFR' , 'SuspensionTravelMetersRL' , 'SuspensionTravelMetersRR' , 'CarClass' , 'CarPerformanceLetterNumber' , 'DrivetrainType' , 'NumberOfCylinders' , 'CarType' , 'ObjectHit' , 'PositionX' , 'PositionY' , 'PositionZ' , 'Speed' , 'Power' , 'Torque' , 'TireTempFL' , 'TireTempFR' , 'TireTempRL' , 'TireTempRR' , 'Boost' , 'Fuel' , 'DistanceTraveled' , 'BestLap' , 'LastLap' , 'CurrentLap' , 'CurrentRaceTime' , 'LapNumber' , 'RacePosition' , 'Throttle' , 'Brake' , 'Clutch' , 'HandBrake' , 'Gear' , 'Steer' , 'NormalizedDrivingLine' , 'NormalizedAIBrakeDifference']
    Titles = ['RaceOn' , 'TimeStamp' , 'MaxRPM' , 'IdleRPM' , 'CurrentRPM' , 'AccelerationX' , 'AccelerationY' , 'AccelerationZ' , 'RezultantAcceleration' , 'VelocityX' , 'VelocityY' , 'VelocityZ' , 'AngularVelocityX' , 'AngularVelocityY' , 'AngularVelocityZ' , 'Yaw' , 'Pitch' , 'Roll' , 'NormalizedSuspensionTravelFL' , 'NormalizedSuspensionTravelFR' , 'NormalizedSuspensionTravelRL' , 'NormalizedSuspensionTravelRR' , 'TireSlipRatioFL' , 'TireSlipRatioFR' , 'TireSlipRatioRL' , 'TireSlipRatioRR' , 'WheelRotationSpeedFL' , 'WheelRotationSpeedFR' , 'WheelRotationSpeedRL' , 'WheelRotationSpeedRR' , 'TireSlipAngleFL' , 'TireSlipAngleFR' , 'TireSlipAngleRL' , 'TireSlipAngleRR' , 'TireCombinedSlipFL' , 'TireCombinedSlipFR' , 'TireCombinedSlipRL' , 'TireCombinedSlipRR' , 'SuspensionTravelMetersFL' , 'SuspensionTravelMetersFR' , 'SuspensionTravelMetersRL' , 'SuspensionTravelMetersRR' , 'CarClass' , 'CarPerformanceLetterNumber' , 'DrivetrainType' , 'NumberOfCylinders' , 'PositionX' , 'PositionY' , 'PositionZ' , 'Speed' , 'Power' , 'Torque' , 'TireTempFL' , 'TireTempFR' , 'TireTempRL' , 'TireTempRR' , 'Boost' , 'Throttle' , 'Brake' , 'Clutch' , 'HandBrake' , 'Gear' , 'Steer' , 'NormalizedDrivingLine' , 'NormalizedAIBrakeDifference']

    for Title in Titles:
        WorkSheet[get_column_letter(LetterNumber) + str(1)].value = Title           # Name each collumn
        LetterNumber += 1

    HostAddress = "127.0.0.1"
    PortNumber = 5300

    Socket = socket.socket(socket.AF_INET, socket.SOCK_RAW, socket.IPPROTO_IP)      # Create socket object
    Socket.bind((HostAddress,PortNumber))                                           # Bind socket
    Socket.setsockopt(socket.IPPROTO_IP, socket.IP_HDRINCL, 1)                      # Include IP headers
    Socket.ioctl(socket.SIO_RCVALL, socket.RCVALL_ON)                               # Receive all packages

    FormatOfHeader = '!BBHHHBBHBBBBBBBB'
    EndOfHeader = 20
    StartOfUsefullDataSegment = 28

    while(True):
        Message = Socket.recvfrom(65565)                                            # The entire received Message
        FullData = Message[0]                                                       # Separate the data from the Address and PortNumber values
        Address = Message[1]                                                        # Separate the Address and PortNumber values from the data field
        Header = struct.unpack(FormatOfHeader, FullData[:EndOfHeader])              # Get only the Header info

        if(Header[6] == 17 and Header[2] == 352):                                   # Filter: 17 -> UDP | 352 -> Total length
            UsefullData = FullData[StartOfUsefullDataSegment:]                      # Separate the useful data from the rest
            
            if FHT.get_race_on(UsefullData) == 'ON':
                ############################## Timer ####################################
                # if FHT.timer(UsefullData, 200) != None:
                #     print(f"{FHT.timer(UsefullData, 200)}")
                #########################################################################

                WorkSheet[get_column_letter(1) + str(DataRow)].value = FHT.get_race_on(UsefullData)
                WorkSheet[get_column_letter(2) + str(DataRow)].value = FHT.get_time_stamp(UsefullData)
                WorkSheet[get_column_letter(3) + str(DataRow)].value = FHT.get_max_rpm(UsefullData)
                WorkSheet[get_column_letter(4) + str(DataRow)].value = FHT.get_idle_rpm(UsefullData)
                WorkSheet[get_column_letter(5) + str(DataRow)].value = FHT.get_rpm(UsefullData)
                WorkSheet[get_column_letter(6) + str(DataRow)].value = FHT.get_AccelX(UsefullData)
                WorkSheet[get_column_letter(7) + str(DataRow)].value = FHT.get_AccelY(UsefullData)
                WorkSheet[get_column_letter(8) + str(DataRow)].value = FHT.get_AccelZ(UsefullData)
                WorkSheet[get_column_letter(9) + str(DataRow)].value = FHT.get_Rez_Accel(UsefullData)
                WorkSheet[get_column_letter(10) + str(DataRow)].value = FHT.get_VelocityX(UsefullData)
                WorkSheet[get_column_letter(11) + str(DataRow)].value = FHT.get_VelocityY(UsefullData)
                WorkSheet[get_column_letter(12) + str(DataRow)].value = FHT.get_VelocityZ(UsefullData)
                WorkSheet[get_column_letter(13) + str(DataRow)].value = FHT.get_angular_VelocityX(UsefullData)
                WorkSheet[get_column_letter(14) + str(DataRow)].value = FHT.get_angular_VelocityY(UsefullData)
                WorkSheet[get_column_letter(15) + str(DataRow)].value = FHT.get_angular_VelocityZ(UsefullData)
                WorkSheet[get_column_letter(16) + str(DataRow)].value = FHT.get_Yaw(UsefullData)
                WorkSheet[get_column_letter(17) + str(DataRow)].value = FHT.get_Pitch(UsefullData)
                WorkSheet[get_column_letter(18) + str(DataRow)].value = FHT.get_Roll(UsefullData)
                WorkSheet[get_column_letter(19) + str(DataRow)].value = FHT.get_normalized_suspension_travel_FL(UsefullData)
                WorkSheet[get_column_letter(20) + str(DataRow)].value = FHT.get_normalized_suspension_travel_FR(UsefullData)
                WorkSheet[get_column_letter(21) + str(DataRow)].value = FHT.get_normalized_suspension_travel_RL(UsefullData)
                WorkSheet[get_column_letter(22) + str(DataRow)].value = FHT.get_normalized_suspension_travel_RR(UsefullData)
                WorkSheet[get_column_letter(23) + str(DataRow)].value = FHT.get_tire_slip_ratio_FL(UsefullData)
                WorkSheet[get_column_letter(24) + str(DataRow)].value = FHT.get_tire_slip_ratio_FR(UsefullData)
                WorkSheet[get_column_letter(25) + str(DataRow)].value = FHT.get_tire_slip_ratio_RL(UsefullData)
                WorkSheet[get_column_letter(26) + str(DataRow)].value = FHT.get_tire_slip_ratio_RR(UsefullData)
                WorkSheet[get_column_letter(27) + str(DataRow)].value = FHT.get_wheel_rot_Speed_FL(UsefullData)
                WorkSheet[get_column_letter(28) + str(DataRow)].value = FHT.get_wheel_rot_Speed_FR(UsefullData)
                WorkSheet[get_column_letter(29) + str(DataRow)].value = FHT.get_wheel_rot_Speed_RL(UsefullData)
                WorkSheet[get_column_letter(30) + str(DataRow)].value = FHT.get_wheel_rot_Speed_RR(UsefullData)
                WorkSheet[get_column_letter(31) + str(DataRow)].value = FHT.get_slip_angle_FL(UsefullData)
                WorkSheet[get_column_letter(32) + str(DataRow)].value = FHT.get_slip_angle_FR(UsefullData)
                WorkSheet[get_column_letter(33) + str(DataRow)].value = FHT.get_slip_angle_RL(UsefullData)
                WorkSheet[get_column_letter(34) + str(DataRow)].value = FHT.get_slip_angle_RR(UsefullData)
                WorkSheet[get_column_letter(35) + str(DataRow)].value = FHT.get_tire_combined_slip_FL(UsefullData)
                WorkSheet[get_column_letter(36) + str(DataRow)].value = FHT.get_tire_combined_slip_FR(UsefullData)
                WorkSheet[get_column_letter(37) + str(DataRow)].value = FHT.get_tire_combined_slip_RL(UsefullData)
                WorkSheet[get_column_letter(38) + str(DataRow)].value = FHT.get_tire_combined_slip_RR(UsefullData)
                WorkSheet[get_column_letter(39) + str(DataRow)].value = FHT.get_suspension_travel_FL(UsefullData)
                WorkSheet[get_column_letter(40) + str(DataRow)].value = FHT.get_suspension_travel_FR(UsefullData)
                WorkSheet[get_column_letter(41) + str(DataRow)].value = FHT.get_suspension_travel_RL(UsefullData)
                WorkSheet[get_column_letter(42) + str(DataRow)].value = FHT.get_suspension_travel_RR(UsefullData)
                WorkSheet[get_column_letter(43) + str(DataRow)].value = FHT.get_car_class(UsefullData)
                WorkSheet[get_column_letter(44) + str(DataRow)].value = FHT.get_performance_index(UsefullData)
                WorkSheet[get_column_letter(45) + str(DataRow)].value = FHT.get_drivetrain_type(UsefullData)
                WorkSheet[get_column_letter(46) + str(DataRow)].value = FHT.get_num_of_cyl(UsefullData)
                WorkSheet[get_column_letter(47) + str(DataRow)].value = FHT.get_CoordX(UsefullData)
                WorkSheet[get_column_letter(48) + str(DataRow)].value = FHT.get_CoordY(UsefullData)
                WorkSheet[get_column_letter(49) + str(DataRow)].value = FHT.get_CoordZ(UsefullData)
                WorkSheet[get_column_letter(50) + str(DataRow)].value = FHT.get_Speed(UsefullData)
                WorkSheet[get_column_letter(51) + str(DataRow)].value = FHT.get_Power(UsefullData)
                WorkSheet[get_column_letter(52) + str(DataRow)].value = FHT.get_Torque(UsefullData)
                WorkSheet[get_column_letter(53) + str(DataRow)].value = FHT.get_Tire_temp_FL(UsefullData)
                WorkSheet[get_column_letter(54) + str(DataRow)].value = FHT.get_Tire_temp_FR(UsefullData)
                WorkSheet[get_column_letter(55) + str(DataRow)].value = FHT.get_Tire_temp_RL(UsefullData)
                WorkSheet[get_column_letter(56) + str(DataRow)].value = FHT.get_Tire_temp_RR(UsefullData)
                WorkSheet[get_column_letter(57) + str(DataRow)].value = FHT.get_Boost(UsefullData)
                WorkSheet[get_column_letter(58) + str(DataRow)].value = FHT.get_Throttle_status(UsefullData)
                WorkSheet[get_column_letter(59) + str(DataRow)].value = FHT.get_Brake_status(UsefullData)
                WorkSheet[get_column_letter(60) + str(DataRow)].value = FHT.get_Clutch_status(UsefullData)
                WorkSheet[get_column_letter(61) + str(DataRow)].value = FHT.get_HandBrake_status(UsefullData)
                WorkSheet[get_column_letter(62) + str(DataRow)].value = FHT.get_Gear(UsefullData)
                WorkSheet[get_column_letter(63) + str(DataRow)].value = FHT.get_Steering_input(UsefullData)
                WorkSheet[get_column_letter(64) + str(DataRow)].value = FHT.get_norm_driving_line(UsefullData)
                WorkSheet[get_column_letter(65) + str(DataRow)].value = FHT.get_norm_AI_brake_diff(UsefullData)
            
                DataRow += 1

            EndTime = int(time.strftime(LogUnit))                                                                              
            print(f'Current time is {EndTime}')

            if int(time.strftime("%S")) % 2:
                try:
                    WorkBook.save(OutputFile)
                    print('Saved...')
                except:
                    print('Error Saving...')

            if EndTime - StartTime >= LogLength :
                break